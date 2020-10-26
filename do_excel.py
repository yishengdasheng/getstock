# -*- coding:utf-8 _*-

#   author : YOYO
#   time :  2020/10/23 16:51
#   email : youyou.xu@enesource.com
#   project_name :  getstock
#   file_name :  do_excel
#   function：操作excel，查询数据，插入数据
from openpyxl import load_workbook

# 存储从excel拿到的数据
class Stock_Data:
    def __init__(self):
        self.code = None
        self.code_name = None
        self.industry = None
        self.code = None
        self.code = None
        self.code = None


class DoExcel:
    def __init__(self):
        self.name = r"E:\桌面\PycharmProjects\getstock\test.xlsx"
        self.sheetname = "Sheet1"

    def open_excel(self):
        excel = load_workbook(self.name)
        sheet = excel[self.sheetname]
        return excel, sheet

    def read(self):
        sheet = self.open_excel()[1]
        datas = []
        for i in range(2, sheet.max_row + 1):
            data = Stock_Data()
            data.code = sheet.cell(i, 2).value
            data.code_name = sheet.cell(i, 3).value
            datas.append(data)
        return datas

    def write_industry(self, row, data):
        excel, sheet = self.open_excel()
        sheet.cell(row, 4).value = data['industry']
        excel.save(self.name)

    def write_k_data(self, row, data):
        excel, sheet = self.open_excel()
        sheet.cell(row, 5).value = data['open']
        sheet.cell(row, 6).value = data['close']
        sheet.cell(row, 7).value = data['high']
        sheet.cell(row, 8).value = data['low']
        sheet.cell(row, 9).value = data['preclose']
        sheet.cell(row, 10).value = data['volume']
        sheet.cell(row, 11).value = data['amount']
        sheet.cell(row, 12).value = data['turn']
        sheet.cell(row, 13).value = data['pctChg']
        sheet.cell(row, 14).value = data['isST']
        excel.save(self.name)